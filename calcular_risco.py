import pandas as pd
import os

ARQUIVO = 'alunos.xlsx'
def calcular_risco(linha) :
    risco = 0
    risco += min(linha['faltas_bimestre'] / 40, 1) * 35
    risco += max(0, (5 - linha['media_notas']) / 5) * 20
    risco += max(0, (3 - linha['renda_salarios']) / 3) * 12
    risco += (linha['reprovacoes'] / 4) * 8
    if linha['trabalha'] == 'Sim':
        risco += 25
    return round(risco)
    
df = pd.read_excel(ARQUIVO, engine='openpyxl')
df['risco'] = df.apply(calcular_risco, axis=1)
df['nivel'] = df['risco'].apply(lambda x: 'ALTO' if x > 55 else 'MODERADO' if x > 25 else 'BAIXO')
print(df[['nome', 'turma', 'risco', 'nivel']].to_string())

writer = pd.ExcelWriter('relatorio_risco.xlsx', engine='openpyxl')
df.to_excel(writer, index=False, sheet_name='Alunos')
workbook = writer.book
sheet = writer.sheets['Alunos']
from openpyxl.styles import PatternFill
verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
amarelo = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
vermelho = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
for i, row in df.iterrows():
    nivel = row['nivel']
    if nivel == 'ALTO':
        cor = vermelho
    elif nivel == 'MODERADO':
        cor = amarelo
    else:
        cor = verde
    for col in range(1, len(df.columns) + 1):
        sheet.cell(row=i+2, column=col).fill = cor
print('Relatório salvo em relatorio_risco.xlsx!')
writer.close()